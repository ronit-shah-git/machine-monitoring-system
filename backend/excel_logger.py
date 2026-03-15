import time
from datetime import datetime

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter

from config import APP_TIMEZONE, EXCEL_IDLE_SLEEP_SECONDS, EXCEL_WRITE_INTERVAL_SECONDS, LOGS_DIR
from state_manager import AppState


class ExcelLogger:
    def __init__(self, state: AppState) -> None:
        self.state = state

    @staticmethod
    def format_timestamp(timestamp: int) -> str:
        return datetime.fromtimestamp(timestamp).strftime("%Y-%m-%d %H:%M:%S")

    @staticmethod
    def format_duration(seconds: int) -> str:
        seconds = max(0, int(seconds))
        hours = seconds // 3600
        minutes = (seconds % 3600) // 60
        seconds = seconds % 60
        return f"{hours:02}:{minutes:02}:{seconds:02}"

    def write_downtimes_to_excel_daily_loop(self) -> None:
        while True:
            try:
                today = datetime.now(APP_TIMEZONE)
                today_str = today.strftime("%Y-%m-%d")
                excel_file = LOGS_DIR / f"downtime_{today_str}.xlsx"

                with self.state.lock:
                    completed_downtimes = [
                        entry for entry in self.state.downtimes if not entry.is_active
                    ]
                    current_count = len(completed_downtimes)

                    if current_count == self.state.last_written_count:
                        time.sleep(EXCEL_IDLE_SLEEP_SECONDS)
                        continue

                    self.state.last_written_count = current_count

                workbook = Workbook()
                sheet = workbook.active
                sheet.title = "Downtime Logs"

                headers = ["Start", "End", "Duration", "Reason", "Active"]
                sheet.append(headers)

                header_fill = PatternFill(
                    start_color="4F81BD",
                    end_color="4F81BD",
                    fill_type="solid",
                )
                header_font = Font(bold=True, color="FFFFFF")

                for col_idx, _ in enumerate(headers, start=1):
                    cell = sheet.cell(row=1, column=col_idx)
                    cell.font = header_font
                    cell.fill = header_fill

                for entry in completed_downtimes:
                    duration = entry.end - entry.start
                    sheet.append(
                        [
                            self.format_timestamp(entry.start),
                            self.format_timestamp(entry.end),
                            self.format_duration(duration),
                            entry.reason,
                            entry.is_active,
                        ]
                    )

                for column in sheet.columns:
                    max_length = max(len(str(cell.value or "")) for cell in column)
                    sheet.column_dimensions[get_column_letter(column[0].column)].width = max(
                        15,
                        max_length + 2,
                    )

                workbook.save(excel_file)
                print(f"✅ Excel written: {excel_file} at {datetime.now(APP_TIMEZONE)}")

            except Exception as exc:
                print(f"❌ Excel write error: {exc}")

            time.sleep(EXCEL_WRITE_INTERVAL_SECONDS)